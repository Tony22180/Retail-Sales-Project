import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference

def create_excel_dashboard():
    # Load data from SQLite
    conn = sqlite3.connect("data/retail.db")
    sales_data = pd.read_sql("SELECT * FROM retail", conn)
    rfm_data = pd.read_excel("reports/rfm_analysis.xlsx")
    conn.close()

    wb = Workbook()
    ws_metrics = wb.active
    ws_metrics.title = "Key Metrics"
    
    metrics = {
        "Total Revenue": sales_data['Price'].sum(),
        "Total Customers": sales_data['Customer_ID'].nunique(),
        "Avg Order Value": sales_data.groupby('Invoice')['Price'].sum().mean()
    }
    for i, (k, v) in enumerate(metrics.items(), start=1):
        ws_metrics.cell(row=i, column=1, value=k)
        ws_metrics.cell(row=i, column=2, value=v)

    ws_country = wb.create_sheet("Sales by Country")
    country_data = sales_data.groupby('Country')['Price'].sum().reset_index()
    for r in dataframe_to_rows(country_data, index=False, header=True):
        ws_country.append(r)
    
    chart = BarChart()
    chart.title = "Revenue by Country"
    chart.x_axis.title = "Country"
    chart.y_axis.title = "Revenue"
    data = Reference(ws_country, min_col=2, min_row=1, max_row=len(country_data)+1)
    cats = Reference(ws_country, min_col=1, min_row=2, max_row=len(country_data)+1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_country.add_chart(chart, "D2")

    wb.save("reports/retail_dashboard.xlsx")

create_excel_dashboard()